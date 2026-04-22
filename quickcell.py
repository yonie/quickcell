#!/usr/bin/env python3
import math
import os
import re
import sys
import threading
from collections import deque
from datetime import date, datetime, time

import gi

gi.require_version("Gtk", "3.0")
gi.require_version("Gdk", "3.0")
gi.require_version("Pango", "1.0")
gi.require_version("PangoCairo", "1.0")
import openpyxl
from gi.repository import Gdk, GLib, Gtk, Pango, PangoCairo
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter

VERSION = "1.1.0"

DEFAULT_COL_WIDTH = 90
DEFAULT_ROW_HEIGHT = 22
HEADER_ROW_HEIGHT = 22
HEADER_COL_WIDTH = 52
CELL_PADDING = 4
MIN_COLS = 12
MIN_ROWS = 30
MIN_ZOOM = 0.3
MAX_ZOOM = 4.0

COLOR_GRID = (0.85, 0.85, 0.85)
COLOR_HEADER_BG = (0.94, 0.94, 0.94)
COLOR_HEADER_BG_ACTIVE = (0.82, 0.88, 0.96)
COLOR_HEADER_TEXT = (0.25, 0.25, 0.25)
COLOR_SEL_FILL = (0.28, 0.52, 0.85, 0.18)
COLOR_SEL_BORDER = (0.18, 0.42, 0.78)
COLOR_CELL_TEXT = (0.1, 0.1, 0.1)
COLOR_CELL_BG = (1.0, 1.0, 1.0)
COLOR_FORMULA_PENDING = (0.55, 0.55, 0.55)
COLOR_FORMULA_PENDING_BG = (0.98, 0.95, 0.80)
COLOR_FORMULA_ERROR = (0.80, 0.20, 0.20)
COLOR_FORMULA_ERROR_BG = (1.00, 0.93, 0.93)


# Default Office 2007+ theme palette (approximation — good enough for viewing)
_DEFAULT_THEME_RGB = [
    "FFFFFF",  # 0 bg1 / light1
    "000000",  # 1 tx1 / dark1
    "E7E6E6",  # 2 bg2 / light2
    "44546A",  # 3 tx2 / dark2
    "4472C4",  # 4 Accent1
    "ED7D31",  # 5 Accent2
    "A5A5A5",  # 6 Accent3
    "FFC000",  # 7 Accent4
    "5B9BD5",  # 8 Accent5
    "70AD47",  # 9 Accent6
    "0563C1",  # 10 Hlink
    "954F72",  # 11 FolHlink
]


def _parse_hex_rgb(s):
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    if len(s) == 8:
        s = s[2:]
    if len(s) != 6:
        return None
    try:
        return (
            int(s[0:2], 16) / 255,
            int(s[2:4], 16) / 255,
            int(s[4:6], 16) / 255,
        )
    except ValueError:
        return None


def _apply_tint(rgb, tint):
    r, g, b = rgb
    if tint < 0:
        f = 1 + tint
        return (max(0, r * f), max(0, g * f), max(0, b * f))
    f = tint
    return (r + (1 - r) * f, g + (1 - g) * f, b + (1 - b) * f)


def color_to_rgb(color):
    """Convert an openpyxl Color to an (r, g, b) tuple in [0,1], or None."""
    if color is None:
        return None
    t = getattr(color, "type", None)
    if t == "rgb":
        rgb = getattr(color, "rgb", None)
        if isinstance(rgb, str):
            return _parse_hex_rgb(rgb)
        return None
    if t == "indexed":
        idx = getattr(color, "indexed", None)
        if idx is None:
            return None
        try:
            return _parse_hex_rgb(COLOR_INDEX[idx])
        except (IndexError, TypeError):
            return None
    if t == "theme":
        idx = getattr(color, "theme", None)
        tint = getattr(color, "tint", 0) or 0
        if idx is None or not (0 <= idx < len(_DEFAULT_THEME_RGB)):
            return None
        base = _parse_hex_rgb(_DEFAULT_THEME_RGB[idx])
        if base is None:
            return None
        return _apply_tint(base, tint) if tint else base
    return None


def _case_insensitive_glob(ext):
    """Turn ``xlsx`` into ``*.[xX][lL][sS][xX]`` so Gtk's case-sensitive
    file filter matches either casing."""
    return "*." + "".join(f"[{c.lower()}{c.upper()}]" for c in ext)


def _fmt_stat(n):
    if n is None:
        return ""
    if isinstance(n, bool):
        return "TRUE" if n else "FALSE"
    try:
        f = float(n)
    except (TypeError, ValueError):
        return str(n)
    if f.is_integer() and abs(f) < 1e16:
        return str(int(f))
    s = f"{f:.6f}".rstrip("0").rstrip(".")
    return s or "0"


def _scroll_deltas(event):
    """Return (dx, dy) from a smooth-scroll event across PyGObject variants."""
    try:
        res = event.get_scroll_deltas()
    except Exception:
        return 0.0, 0.0
    if res is None:
        return 0.0, 0.0
    if len(res) == 3:
        _, dx, dy = res
    elif len(res) == 2:
        dx, dy = res
    else:
        return 0.0, 0.0
    return float(dx or 0.0), float(dy or 0.0)


def format_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if value.hour or value.minute or value.second:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float):
        if value.is_integer() and abs(value) < 1e16:
            return str(int(value))
        return f"{value:.10g}"
    return str(value)


# ---------------------------------------------------------------------------
# Minimal same-sheet formula evaluator.
# Handles arithmetic, comparisons, string concat, cell refs, A1:B2 ranges,
# and a small whitelist of common functions. Anything unsupported raises
# FormulaError, which the caller surfaces as "Error" in the cell.
# ---------------------------------------------------------------------------


class FormulaError(Exception):
    pass


class FormulaErrorValue:
    __slots__ = ("msg",)

    def __init__(self, msg):
        self.msg = msg


FORMULA_PENDING = object()


_CELL_REF_RE = re.compile(r"\$?([A-Za-z]+)\$?(\d+)")
_COL_ONLY_RE = re.compile(r"\$?([A-Za-z]+)")
_ROW_ONLY_RE = re.compile(r"\$?(\d+)")


def _col_letters_to_num(letters):
    n = 0
    for c in letters.upper():
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n


def _parse_cell_ref(name):
    m = _CELL_REF_RE.fullmatch(name)
    if not m:
        return None
    return int(m.group(2)), _col_letters_to_num(m.group(1))


def _parse_col_only(name):
    m = _COL_ONLY_RE.fullmatch(name)
    if not m:
        return None
    return _col_letters_to_num(m.group(1))


def _parse_row_only(name):
    m = _ROW_ONLY_RE.fullmatch(name)
    if not m:
        return None
    return int(m.group(1))


def _to_num(v):
    if isinstance(v, bool):
        return 1 if v else 0
    if isinstance(v, (int, float)):
        return v
    if v is None or v == "":
        return 0
    if isinstance(v, str):
        try:
            return float(v) if any(c in v for c in ".eE") else int(v)
        except ValueError:
            raise FormulaError(f"not a number: {v!r}")
    raise FormulaError(f"not a number: {type(v).__name__}")


def _to_str(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        if v.is_integer() and abs(v) < 1e16:
            return str(int(v))
        return f"{v:.10g}"
    return str(v)


def _truthy(v):
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    if isinstance(v, str):
        return v.upper() == "TRUE"
    return bool(v)


def _flatten(v, out=None):
    if out is None:
        out = []
    if isinstance(v, list):
        for item in v:
            _flatten(item, out)
    else:
        out.append(v)
    return out


def _numbers_only(vs):
    return [v for v in vs if isinstance(v, (int, float)) and not isinstance(v, bool)]


def _call_func(name, args):
    flat = _flatten(args)
    if name == "SUM":
        return sum(_numbers_only(flat))
    if name in ("AVG", "AVERAGE"):
        nums = _numbers_only(flat)
        if not nums:
            raise FormulaError("AVERAGE of empty set")
        return sum(nums) / len(nums)
    if name == "MIN":
        nums = _numbers_only(flat)
        if not nums:
            raise FormulaError("MIN of empty set")
        return min(nums)
    if name == "MAX":
        nums = _numbers_only(flat)
        if not nums:
            raise FormulaError("MAX of empty set")
        return max(nums)
    if name == "COUNT":
        return len(_numbers_only(flat))
    if name == "COUNTA":
        return sum(1 for v in flat if v is not None and v != "")
    if name == "COUNTIF":
        if len(args) != 2:
            raise FormulaError("COUNTIF takes 2 args")
        return sum(1 for v in _flatten(args[0]) if _matches_criterion(v, args[1]))
    if name == "SUMIF":
        if len(args) not in (2, 3):
            raise FormulaError("SUMIF takes 2-3 args")
        rng = _flatten(args[0])
        crit = args[1]
        sum_range = _flatten(args[2]) if len(args) == 3 else rng
        total = 0.0
        for i, v in enumerate(rng):
            if _matches_criterion(v, crit):
                if i < len(sum_range):
                    sv = sum_range[i]
                    if isinstance(sv, (int, float)) and not isinstance(sv, bool):
                        total += sv
        return total
    if name == "PRODUCT":
        nums = _numbers_only(flat)
        r = 1.0
        for n in nums:
            r *= n
        return r
    if name == "ABS":
        if len(args) != 1:
            raise FormulaError("ABS takes 1 arg")
        return abs(_to_num(args[0]))
    if name == "ROUND":
        if len(args) != 2:
            raise FormulaError("ROUND takes 2 args")
        return round(_to_num(args[0]), int(_to_num(args[1])))
    if name == "INT":
        if len(args) != 1:
            raise FormulaError("INT takes 1 arg")
        return math.floor(_to_num(args[0]))
    if name == "MOD":
        if len(args) != 2:
            raise FormulaError("MOD takes 2 args")
        d = _to_num(args[1])
        if d == 0:
            raise FormulaError("MOD divide by zero")
        return _to_num(args[0]) % d
    if name == "SQRT":
        if len(args) != 1:
            raise FormulaError("SQRT takes 1 arg")
        x = _to_num(args[0])
        if x < 0:
            raise FormulaError("SQRT of negative")
        return math.sqrt(x)
    if name == "POWER":
        if len(args) != 2:
            raise FormulaError("POWER takes 2 args")
        return _to_num(args[0]) ** _to_num(args[1])
    if name == "IF":
        if len(args) not in (2, 3):
            raise FormulaError("IF takes 2-3 args")
        if _truthy(args[0]):
            return args[1]
        return args[2] if len(args) == 3 else False
    if name == "AND":
        return all(_truthy(v) for v in flat)
    if name == "OR":
        return any(_truthy(v) for v in flat)
    if name == "NOT":
        if len(args) != 1:
            raise FormulaError("NOT takes 1 arg")
        return not _truthy(args[0])
    if name in ("CONCAT", "CONCATENATE"):
        return "".join(_to_str(v) for v in flat)
    if name == "LEN":
        if len(args) != 1:
            raise FormulaError("LEN takes 1 arg")
        return len(_to_str(args[0]))
    if name == "LEFT":
        if len(args) == 1:
            n = 1
        elif len(args) == 2:
            n = int(_to_num(args[1]))
        else:
            raise FormulaError("LEFT takes 1-2 args")
        return _to_str(args[0])[: max(0, n)]
    if name == "RIGHT":
        if len(args) == 1:
            n = 1
        elif len(args) == 2:
            n = int(_to_num(args[1]))
        else:
            raise FormulaError("RIGHT takes 1-2 args")
        s = _to_str(args[0])
        return s[-n:] if n > 0 else ""
    if name == "MID":
        if len(args) != 3:
            raise FormulaError("MID takes 3 args")
        s = _to_str(args[0])
        start = max(1, int(_to_num(args[1])))
        length = max(0, int(_to_num(args[2])))
        return s[start - 1 : start - 1 + length]
    if name == "UPPER":
        if len(args) != 1:
            raise FormulaError("UPPER takes 1 arg")
        return _to_str(args[0]).upper()
    if name == "LOWER":
        if len(args) != 1:
            raise FormulaError("LOWER takes 1 arg")
        return _to_str(args[0]).lower()
    if name == "TRIM":
        if len(args) != 1:
            raise FormulaError("TRIM takes 1 arg")
        return " ".join(_to_str(args[0]).split())
    if name == "VALUE":
        if len(args) != 1:
            raise FormulaError("VALUE takes 1 arg")
        return _to_num(args[0])
    if name == "INDEX":
        if len(args) == 2:
            arr = _flatten(args[0])
            n = int(_to_num(args[1]))
            if n < 1 or n > len(arr):
                raise FormulaError("INDEX out of range")
            return arr[n - 1]
        raise FormulaError("INDEX 2D not supported")
    if name == "MATCH":
        if len(args) not in (2, 3):
            raise FormulaError("MATCH takes 2-3 args")
        lookup = args[0]
        arr = _flatten(args[1])
        match_type = 1 if len(args) < 3 else int(_to_num(args[2]))
        if match_type == 0:
            for i, v in enumerate(arr, 1):
                if _cmp_eq(v, lookup):
                    return i
            raise FormulaError("no match")
        if match_type == 1:
            target = _to_num(lookup)
            best = None
            for i, v in enumerate(arr, 1):
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    if v <= target:
                        best = i
                    else:
                        break
            if best is None:
                raise FormulaError("no match")
            return best
        raise FormulaError("MATCH type -1 not supported")
    raise FormulaError(f"unsupported function: {name}")


def _wildcard_to_regex(s):
    out = []
    i = 0
    while i < len(s):
        c = s[i]
        if c == "~" and i + 1 < len(s) and s[i + 1] in "*?~":
            out.append(re.escape(s[i + 1]))
            i += 2
            continue
        if c == "*":
            out.append(".*")
        elif c == "?":
            out.append(".")
        else:
            out.append(re.escape(c))
        i += 1
    return re.compile("".join(out), re.IGNORECASE | re.DOTALL)


def _matches_criterion(value, criterion):
    if isinstance(criterion, str):
        s = criterion
        op = None
        rest = ""
        for prefix in (">=", "<=", "<>", ">", "<", "="):
            if s.startswith(prefix):
                op = prefix
                rest = s[len(prefix) :]
                break
        if op is not None:
            try:
                target = float(rest) if any(c in rest for c in ".eE") else int(rest)
            except ValueError:
                target = rest
            if op == "=":
                return _cmp_eq(value, target)
            if op == "<>":
                return not _cmp_eq(value, target)
            try:
                v = _to_num(value)
                t = _to_num(target)
            except FormulaError:
                return False
            if op == ">":
                return v > t
            if op == "<":
                return v < t
            if op == ">=":
                return v >= t
            if op == "<=":
                return v <= t
        if "*" in s or "?" in s:
            return bool(_wildcard_to_regex(s).fullmatch(_to_str(value)))
    return _cmp_eq(value, criterion)


def _cmp_eq(a, b):
    if (
        isinstance(a, (int, float))
        and isinstance(b, (int, float))
        and not isinstance(a, bool)
        and not isinstance(b, bool)
    ):
        return a == b
    return _to_str(a).upper() == _to_str(b).upper()


class _FormulaParser:
    """Recursive-descent parser for a single formula. Evaluates as it parses.

    `ctx` must implement:
        get_value(sheet_name, row, col) -> scalar, raises FormulaError
        get_sheet_dims(sheet_name) -> (max_row, max_col) or None

    `sheet_name` is the sheet that unqualified cell refs resolve against.
    """

    def __init__(self, src, ctx, sheet_name):
        self.src = src
        self.pos = 0
        self.ctx = ctx
        self.sheet_name = sheet_name

    def parse(self):
        if not self.src.startswith("="):
            raise FormulaError("not a formula")
        self.pos = 0
        while self.pos < len(self.src) and self.src[self.pos] == "=":
            self.pos += 1
        v = self._expr()
        self._ws()
        if self.pos < len(self.src):
            raise FormulaError(f"trailing text: {self.src[self.pos :]!r}")
        return v

    def _ws(self):
        while self.pos < len(self.src) and self.src[self.pos] in " \t\n":
            self.pos += 1

    def _at(self, s):
        self._ws()
        return self.src.startswith(s, self.pos)

    def _consume(self, s):
        if self._at(s):
            self.pos += len(s)
            return True
        return False

    def _expect(self, ch):
        if not self._consume(ch):
            raise FormulaError(f"expected {ch!r}")

    def _expr(self):
        return self._compare()

    def _compare(self):
        left = self._concat()
        while True:
            if self._consume("<="):
                left = _to_num(left) <= _to_num(self._concat())
            elif self._consume(">="):
                left = _to_num(left) >= _to_num(self._concat())
            elif self._consume("<>"):
                left = not _cmp_eq(left, self._concat())
            elif self._consume("<"):
                left = _to_num(left) < _to_num(self._concat())
            elif self._consume(">"):
                left = _to_num(left) > _to_num(self._concat())
            elif self._consume("="):
                left = _cmp_eq(left, self._concat())
            else:
                return left

    def _concat(self):
        left = self._add()
        while self._consume("&"):
            left = _to_str(left) + _to_str(self._add())
        return left

    def _add(self):
        left = self._mul()
        while True:
            if self._consume("+"):
                left = _to_num(left) + _to_num(self._mul())
            elif self._consume("-"):
                left = _to_num(left) - _to_num(self._mul())
            else:
                return left

    def _mul(self):
        left = self._pow()
        while True:
            if self._consume("*"):
                left = _to_num(left) * _to_num(self._pow())
            elif self._consume("/"):
                r = _to_num(self._pow())
                if r == 0:
                    raise FormulaError("divide by zero")
                left = _to_num(left) / r
            else:
                return left

    def _pow(self):
        left = self._unary()
        if self._consume("^"):
            left = _to_num(left) ** _to_num(self._unary())
        return left

    def _unary(self):
        if self._consume("-"):
            return -_to_num(self._unary())
        if self._consume("+"):
            return _to_num(self._unary())
        return self._atom()

    def _atom(self):
        self._ws()
        if self.pos >= len(self.src):
            raise FormulaError("unexpected end")
        ch = self.src[self.pos]
        if ch == "(":
            self.pos += 1
            v = self._expr()
            self._expect(")")
            return v
        if ch == '"':
            return self._string()
        if ch == "'":
            return self._quoted_sheet_ref()
        if ch.isdigit() or ch == ".":
            return self._number()
        if ch.isalpha() or ch == "_" or ch == "$":
            return self._name_or_ref()
        raise FormulaError(f"unexpected {ch!r}")

    def _quoted_sheet_ref(self):
        self.pos += 1  # opening '
        start = self.pos
        while self.pos < len(self.src) and self.src[self.pos] != "'":
            self.pos += 1
        if self.pos >= len(self.src):
            raise FormulaError("unterminated sheet name")
        sheet = self.src[start : self.pos]
        self.pos += 1  # closing '
        if not self._consume("!"):
            raise FormulaError("expected '!' after sheet name")
        return self._parse_ref_on_sheet(sheet)

    def _read_name(self):
        start = self.pos
        while self.pos < len(self.src):
            c = self.src[self.pos]
            if c.isalnum() or c in "$_.":
                self.pos += 1
            else:
                break
        return self.src[start : self.pos]

    def _parse_ref_on_sheet(self, sheet):
        name = self._read_name()
        if not name:
            raise FormulaError(f"expected ref after {sheet}!")
        if self._consume(":"):
            name2 = self._read_name()
            if not name2:
                raise FormulaError(f"expected ref after {sheet}!{name}:")
            return self._resolve_range(sheet, name, name2)
        ref = _parse_cell_ref(name)
        if ref is not None:
            return self.ctx.get_value(sheet, ref[0], ref[1])
        raise FormulaError(f"bad ref {sheet}!{name}")

    def _string(self):
        self.pos += 1
        out = []
        while self.pos < len(self.src):
            c = self.src[self.pos]
            if c == '"':
                if self.pos + 1 < len(self.src) and self.src[self.pos + 1] == '"':
                    out.append('"')
                    self.pos += 2
                    continue
                self.pos += 1
                return "".join(out)
            out.append(c)
            self.pos += 1
        raise FormulaError("unterminated string")

    def _number(self):
        start = self.pos
        while self.pos < len(self.src) and (
            self.src[self.pos].isdigit() or self.src[self.pos] == "."
        ):
            self.pos += 1
        if self.pos < len(self.src) and self.src[self.pos] in "eE":
            self.pos += 1
            if self.pos < len(self.src) and self.src[self.pos] in "+-":
                self.pos += 1
            while self.pos < len(self.src) and self.src[self.pos].isdigit():
                self.pos += 1
        s = self.src[start : self.pos]
        try:
            return float(s) if any(c in s for c in ".eE") else int(s)
        except ValueError:
            raise FormulaError(f"bad number {s!r}")

    def _name_or_ref(self):
        name = self._read_name()
        upper = name.upper()
        if self._at("!"):
            self.pos += 1
            return self._parse_ref_on_sheet(name)
        if self._at("("):
            if upper == "IF":
                return self._call_if()
            if upper == "IFERROR":
                return self._call_iferror()
            self.pos += 1
            args = []
            if not self._at(")"):
                args.append(self._expr())
                while self._consume(","):
                    args.append(self._expr())
            self._expect(")")
            return _call_func(upper, args)
        if self._consume(":"):
            name2 = self._read_name()
            return self._resolve_range(self.sheet_name, name, name2)
        if upper == "TRUE":
            return True
        if upper == "FALSE":
            return False
        ref = _parse_cell_ref(name)
        if ref is not None:
            return self.ctx.get_value(self.sheet_name, ref[0], ref[1])
        raise FormulaError(f"bad name {name!r}")

    def _resolve_range(self, sheet, a, b):
        p1 = _parse_cell_ref(a)
        p2 = _parse_cell_ref(b)
        if p1 and p2:
            r1, c1 = p1
            r2, c2 = p2
        else:
            c1 = _parse_col_only(a)
            c2 = _parse_col_only(b)
            if c1 is not None and c2 is not None:
                dims = self.ctx.get_sheet_dims(sheet)
                if dims is None:
                    raise FormulaError(f"unknown sheet: {sheet}")
                r1, r2 = 1, dims[0]
            else:
                r1 = _parse_row_only(a)
                r2 = _parse_row_only(b)
                if r1 is None or r2 is None:
                    raise FormulaError(f"bad range {a}:{b}")
                dims = self.ctx.get_sheet_dims(sheet)
                if dims is None:
                    raise FormulaError(f"unknown sheet: {sheet}")
                c1, c2 = 1, dims[1]
        rmin, rmax = min(r1, r2), max(r1, r2)
        cmin, cmax = min(c1, c2), max(c1, c2)
        if (rmax - rmin + 1) * (cmax - cmin + 1) > 200_000:
            raise FormulaError("range too large")
        out = []
        for r in range(rmin, rmax + 1):
            for c in range(cmin, cmax + 1):
                out.append(self.ctx.get_value(sheet, r, c))
        return out

    def _call_if(self):
        self.pos += 1  # (
        cond = self._expr()
        self._expect(",")
        if _truthy(cond):
            val = self._expr()
            if self._consume(","):
                self._skip_arg()
        else:
            self._skip_arg()
            if self._consume(","):
                val = self._expr()
            else:
                val = False
        self._expect(")")
        return val

    def _call_iferror(self):
        self.pos += 1  # (
        arg1_start = self.pos
        try:
            val = self._expr()
            self._expect(",")
            self._skip_arg()
        except FormulaError:
            self.pos = arg1_start
            self._skip_arg()
            self._expect(",")
            val = self._expr()
        self._expect(")")
        return val

    def _skip_arg(self):
        depth = 0
        while self.pos < len(self.src):
            c = self.src[self.pos]
            if c == '"':
                self.pos += 1
                while self.pos < len(self.src):
                    if self.src[self.pos] == '"':
                        if (
                            self.pos + 1 < len(self.src)
                            and self.src[self.pos + 1] == '"'
                        ):
                            self.pos += 2
                            continue
                        self.pos += 1
                        break
                    self.pos += 1
                continue
            if c == "'":
                self.pos += 1
                while self.pos < len(self.src):
                    if self.src[self.pos] == "'":
                        self.pos += 1
                        break
                    self.pos += 1
                continue
            if c == "(":
                depth += 1
                self.pos += 1
                continue
            if c == ")":
                if depth == 0:
                    return
                depth -= 1
                self.pos += 1
                continue
            if c == "," and depth == 0:
                return
            self.pos += 1


class WorkbookEvaluator:
    """Workbook-wide formula evaluator: single worker thread, shared cycle
    detection, and access to every sheet's values + formulas so cross-sheet
    refs work."""

    def __init__(self):
        self._sheets = {}  # sheet_name -> SheetView
        self._pending_deque = deque()
        self._pending_set = set()
        self._cv = threading.Condition()
        self._eval_stack = set()
        self._worker_started = False
        self._redraw_pending = {}  # SheetView -> bool

    def register(self, name, sheet_view):
        self._sheets[name] = sheet_view

    def enqueue(self, sheet_view, row, col):
        key = (sheet_view, row, col)
        with self._cv:
            if key in self._pending_set:
                try:
                    self._pending_deque.remove(key)
                except ValueError:
                    pass
            else:
                self._pending_set.add(key)
            self._pending_deque.append(key)
            self._cv.notify()
        self._ensure_worker()

    def _ensure_worker(self):
        if self._worker_started:
            return
        self._worker_started = True
        t = threading.Thread(target=self._worker_loop, daemon=True)
        t.start()

    def _worker_loop(self):
        while True:
            with self._cv:
                while not self._pending_deque:
                    self._cv.wait()
                item = self._pending_deque.pop()
                self._pending_set.discard(item)
            sv, r, c = item
            if (r, c) in sv.formula_cache:
                continue
            self._eval_stack.clear()
            try:
                self.get_value(sv.sheet_name, r, c)
            except FormulaError:
                pass
            except Exception as e:
                sv.formula_cache[(r, c)] = FormulaErrorValue(f"internal: {e}")
            self._schedule_redraw(sv)

    def _schedule_redraw(self, sv):
        if self._redraw_pending.get(sv):
            return
        self._redraw_pending[sv] = True
        GLib.idle_add(self._do_redraw, sv)

    def _do_redraw(self, sv):
        self._redraw_pending[sv] = False
        sv.drawing_area.queue_draw()
        sv._notify_selection()
        return False

    def get_value(self, sheet_name, row, col):
        sv = self._sheets.get(sheet_name)
        if sv is None:
            raise FormulaError(f"unknown sheet: {sheet_name!r}")
        cached = sv.values.get((row, col))
        if cached is not None:
            return cached
        f = sv.formulas.get((row, col))
        if f is None:
            return None
        if (row, col) in sv.formula_cache:
            v = sv.formula_cache[(row, col)]
            if isinstance(v, FormulaErrorValue):
                raise FormulaError(v.msg)
            return v
        key = (sheet_name, row, col)
        if key in self._eval_stack:
            raise FormulaError("circular reference")
        self._eval_stack.add(key)
        try:
            v = _FormulaParser(f, self, sheet_name).parse()
        except FormulaError as e:
            sv.formula_cache[(row, col)] = FormulaErrorValue(str(e))
            raise
        except Exception as e:
            sv.formula_cache[(row, col)] = FormulaErrorValue(f"internal: {e}")
            raise FormulaError(f"internal: {e}")
        finally:
            self._eval_stack.discard(key)
        sv.formula_cache[(row, col)] = v
        return v

    def get_sheet_dims(self, sheet_name):
        sv = self._sheets.get(sheet_name)
        if sv is None:
            return None
        return (sv.max_row, sv.max_col)


class SheetView(Gtk.Grid):
    def __init__(
        self,
        worksheet,
        sheet_name=None,
        ws_formulas=None,
        evaluator=None,
        prebuilt_values=None,
        prebuilt_formulas=None,
        on_selection_changed=None,
        on_zoom_changed=None,
    ):
        super().__init__()
        self.ws = worksheet
        self.ws_formulas = ws_formulas
        self.sheet_name = sheet_name or getattr(worksheet, "title", None) or ""
        self.evaluator = evaluator
        self.on_selection_changed = on_selection_changed
        self.on_zoom_changed = on_zoom_changed

        self.zoom = 1.0

        self.max_row = max(worksheet.max_row or 0, MIN_ROWS)
        self.max_col = max(worksheet.max_column or 0, MIN_COLS)

        if prebuilt_values is not None:
            self.values = prebuilt_values
        else:
            self.values = {}
            for row in worksheet.iter_rows():
                for cell in row:
                    v = cell.value
                    if v is not None:
                        self.values[(cell.row, cell.column)] = v

        if prebuilt_formulas is not None:
            self.formulas = prebuilt_formulas
        else:
            self.formulas = {}
            if ws_formulas is not None:
                for row in ws_formulas.iter_rows():
                    for cell in row:
                        v = cell.value
                        if isinstance(v, str) and v.startswith("="):
                            self.formulas[(cell.row, cell.column)] = v

        self.formula_cache = {}

        if self.evaluator is not None:
            self.evaluator.register(self.sheet_name, self)

        self.base_col_widths = []
        for i in range(1, self.max_col + 1):
            letter = get_column_letter(i)
            dim = worksheet.column_dimensions.get(letter)
            if dim and dim.width:
                px = int(dim.width * 7.0) + 5
                self.base_col_widths.append(max(20, px))
            else:
                self.base_col_widths.append(DEFAULT_COL_WIDTH)

        self.base_row_heights = []
        for i in range(1, self.max_row + 1):
            dim = worksheet.row_dimensions.get(i)
            if dim and dim.height:
                px = int(dim.height * 1.333)
                self.base_row_heights.append(max(14, px))
            else:
                self.base_row_heights.append(DEFAULT_ROW_HEIGHT)

        self.merged_anchor = {}
        self.merged_covered = {}
        for mr in worksheet.merged_cells.ranges:
            top, left = mr.min_row, mr.min_col
            bottom, right = mr.max_row, mr.max_col
            nrows = bottom - top + 1
            ncols = right - left + 1
            self.merged_anchor[(top, left)] = (nrows, ncols)
            for r in range(top, bottom + 1):
                for c in range(left, right + 1):
                    if (r, c) != (top, left):
                        self.merged_covered[(r, c)] = (top, left)

        self.sel_anchor = (1, 1)
        self.sel_cursor = (1, 1)
        self.dragging_sel = False
        self._drag_axis = "cell"

        self.scroll_x = 0
        self.scroll_y = 0

        self.drawing_area = Gtk.DrawingArea()
        self.drawing_area.set_can_focus(True)
        self.drawing_area.set_hexpand(True)
        self.drawing_area.set_vexpand(True)
        self.drawing_area.add_events(
            Gdk.EventMask.BUTTON_PRESS_MASK
            | Gdk.EventMask.BUTTON_RELEASE_MASK
            | Gdk.EventMask.POINTER_MOTION_MASK
            | Gdk.EventMask.SCROLL_MASK
            | Gdk.EventMask.SMOOTH_SCROLL_MASK
            | Gdk.EventMask.KEY_PRESS_MASK
            | Gdk.EventMask.FOCUS_CHANGE_MASK
        )
        self.drawing_area.connect("draw", self._on_draw)
        self.drawing_area.connect("button-press-event", self._on_button_press)
        self.drawing_area.connect("motion-notify-event", self._on_motion)
        self.drawing_area.connect("button-release-event", self._on_button_release)
        self.drawing_area.connect("scroll-event", self._on_scroll)
        self.drawing_area.connect("key-press-event", self._on_key_press)
        self.drawing_area.connect("size-allocate", self._on_size_allocate)

        self.hadj = Gtk.Adjustment()
        self.hadj.set_step_increment(20)
        self.hadj.set_page_increment(200)
        self.vadj = Gtk.Adjustment()
        self.vadj.set_step_increment(20)
        self.vadj.set_page_increment(200)
        self.hadj.connect("value-changed", self._on_hadj)
        self.vadj.connect("value-changed", self._on_vadj)
        self.hscroll = Gtk.Scrollbar(
            orientation=Gtk.Orientation.HORIZONTAL, adjustment=self.hadj
        )
        self.vscroll = Gtk.Scrollbar(
            orientation=Gtk.Orientation.VERTICAL, adjustment=self.vadj
        )

        self.attach(self.drawing_area, 0, 0, 1, 1)
        self.attach(self.vscroll, 1, 0, 1, 1)
        self.attach(self.hscroll, 0, 1, 1, 1)

    @property
    def col_widths(self):
        return [max(8, int(w * self.zoom)) for w in self.base_col_widths]

    @property
    def row_heights(self):
        return [max(8, int(h * self.zoom)) for h in self.base_row_heights]

    @property
    def header_row_height(self):
        return max(14, int(HEADER_ROW_HEIGHT * self.zoom))

    @property
    def header_col_width(self):
        return max(26, int(HEADER_COL_WIDTH * self.zoom))

    @property
    def content_width(self):
        return self.header_col_width + sum(self.col_widths)

    @property
    def content_height(self):
        return self.header_row_height + sum(self.row_heights)

    def _on_size_allocate(self, widget, rect):
        self._update_adjustments()

    def _update_adjustments(self):
        alloc = self.drawing_area.get_allocation()
        vw = max(1, alloc.width)
        vh = max(1, alloc.height)
        self.hadj.configure(
            self.hadj.get_value(),
            0,
            max(vw, self.content_width),
            20,
            max(1, int(vw * 0.8)),
            vw,
        )
        self.vadj.configure(
            self.vadj.get_value(),
            0,
            max(vh, self.content_height),
            20,
            max(1, int(vh * 0.8)),
            vh,
        )
        self.scroll_x = int(self.hadj.get_value())
        self.scroll_y = int(self.vadj.get_value())

    def _on_hadj(self, adj):
        self.scroll_x = int(adj.get_value())
        self.drawing_area.queue_draw()

    def _on_vadj(self, adj):
        self.scroll_y = int(adj.get_value())
        self.drawing_area.queue_draw()

    def _col_x(self, col_1based):
        x = self.header_col_width
        widths = self.col_widths
        for i in range(col_1based - 1):
            x += widths[i]
        return x

    def _row_y(self, row_1based):
        y = self.header_row_height
        heights = self.row_heights
        for i in range(row_1based - 1):
            y += heights[i]
        return y

    def _col_at_x(self, x_doc):
        if x_doc < self.header_col_width:
            return 0
        cx = self.header_col_width
        for i, w in enumerate(self.col_widths):
            if cx + w > x_doc:
                return i + 1
            cx += w
        return len(self.col_widths)

    def _row_at_y(self, y_doc):
        if y_doc < self.header_row_height:
            return 0
        cy = self.header_row_height
        for i, h in enumerate(self.row_heights):
            if cy + h > y_doc:
                return i + 1
            cy += h
        return len(self.row_heights)

    def _on_draw(self, widget, cr):
        alloc = widget.get_allocation()
        vw, vh = alloc.width, alloc.height

        cr.set_source_rgb(*COLOR_CELL_BG)
        cr.rectangle(0, 0, vw, vh)
        cr.fill()

        cws = self.col_widths
        rhs = self.row_heights
        hrh = self.header_row_height
        hcw = self.header_col_width

        first_col = last_col = None
        x = hcw
        for i, w in enumerate(cws):
            if x + w > self.scroll_x + hcw and x < self.scroll_x + vw:
                if first_col is None:
                    first_col = i + 1
                last_col = i + 1
            x += w
        first_row = last_row = None
        y = hrh
        for i, h in enumerate(rhs):
            if y + h > self.scroll_y + hrh and y < self.scroll_y + vh:
                if first_row is None:
                    first_row = i + 1
                last_row = i + 1
            y += h
        if first_col is None or first_row is None:
            return False

        header_font = Pango.FontDescription.new()
        header_font.set_family("Sans")
        header_font.set_size(int(9 * self.zoom * Pango.SCALE))

        sel_t, sel_l, sel_b, sel_r = self._sel_rect()

        cr.save()
        cr.rectangle(hcw, hrh, vw - hcw, vh - hrh)
        cr.clip()

        cr.set_source_rgb(*COLOR_GRID)
        cr.set_line_width(1)
        for row in range(first_row, last_row + 2):
            if row - 1 >= len(rhs):
                break
            y_doc = self._row_y(row)
            sy = y_doc - self.scroll_y
            cr.move_to(hcw, sy + 0.5)
            cr.line_to(vw, sy + 0.5)
            cr.stroke()
        for col in range(first_col, last_col + 2):
            if col - 1 >= len(cws):
                break
            x_doc = self._col_x(col)
            sx = x_doc - self.scroll_x
            cr.move_to(sx + 0.5, hrh)
            cr.line_to(sx + 0.5, vh)
            cr.stroke()

        for row in range(first_row, last_row + 1):
            for col in range(first_col, last_col + 1):
                if (row, col) in self.merged_covered:
                    continue
                anchor = self.merged_anchor.get((row, col))
                nrows, ncols = anchor if anchor else (1, 1)

                x_doc = self._col_x(col)
                y_doc = self._row_y(row)
                w_doc = sum(cws[col - 1 : col - 1 + ncols])
                h_doc = sum(rhs[row - 1 : row - 1 + nrows])
                sx = x_doc - self.scroll_x
                sy = y_doc - self.scroll_y

                cell = self.ws.cell(row=row, column=col)
                display_val = self.resolve_display_value(row, col)

                bg_rgb = None
                fill = cell.fill
                if fill is not None and getattr(fill, "patternType", None) == "solid":
                    bg_rgb = color_to_rgb(fill.fgColor) or color_to_rgb(
                        fill.start_color
                    )
                if display_val is FORMULA_PENDING and bg_rgb is None:
                    bg_rgb = COLOR_FORMULA_PENDING_BG
                elif isinstance(display_val, FormulaErrorValue) and bg_rgb is None:
                    bg_rgb = COLOR_FORMULA_ERROR_BG
                if bg_rgb is not None:
                    cr.set_source_rgb(*bg_rgb)
                    cr.rectangle(sx + 1, sy + 1, w_doc - 1, h_doc - 1)
                    cr.fill()
                elif nrows > 1 or ncols > 1:
                    cr.set_source_rgb(*COLOR_CELL_BG)
                    cr.rectangle(sx + 1, sy + 1, w_doc - 1, h_doc - 1)
                    cr.fill()

                if nrows > 1 or ncols > 1:
                    cr.set_source_rgb(*COLOR_GRID)
                    cr.set_line_width(1)
                    cr.rectangle(sx + 0.5, sy + 0.5, w_doc, h_doc)
                    cr.stroke()

                if sel_t <= row <= sel_b and sel_l <= col <= sel_r:
                    cr.set_source_rgba(*COLOR_SEL_FILL)
                    cr.rectangle(sx, sy, w_doc, h_doc)
                    cr.fill()

                if display_val is FORMULA_PENDING:
                    text = "…"
                    text_override_color = COLOR_FORMULA_PENDING
                elif isinstance(display_val, FormulaErrorValue):
                    text = "Error"
                    text_override_color = COLOR_FORMULA_ERROR
                else:
                    text = format_cell_value(display_val)
                    text_override_color = None
                if text:
                    font = cell.font
                    family = font.name if font and font.name else "Sans"
                    size_pt = (font.size if font and font.size else 10) * self.zoom
                    bold = bool(font and font.bold)
                    italic = bool(font and font.italic)
                    fg_rgb = color_to_rgb(font.color) if font else None

                    font_desc = Pango.FontDescription.new()
                    font_desc.set_family(family)
                    font_desc.set_size(int(max(1, size_pt) * Pango.SCALE))
                    if bold:
                        font_desc.set_weight(Pango.Weight.BOLD)
                    if italic:
                        font_desc.set_style(Pango.Style.ITALIC)

                    cr.set_source_rgb(
                        *(text_override_color or fg_rgb or COLOR_CELL_TEXT)
                    )
                    layout = PangoCairo.create_layout(cr)
                    layout.set_font_description(font_desc)
                    layout.set_text(text, -1)
                    avail = max(1, w_doc - 2 * CELL_PADDING)
                    layout.set_width(avail * Pango.SCALE)
                    layout.set_ellipsize(Pango.EllipsizeMode.END)

                    align_val = (
                        display_val
                        if not (
                            display_val is FORMULA_PENDING
                            or isinstance(display_val, FormulaErrorValue)
                        )
                        else None
                    )
                    align = getattr(cell.alignment, "horizontal", None)
                    if align == "left":
                        pango_align = Pango.Alignment.LEFT
                    elif align == "right":
                        pango_align = Pango.Alignment.RIGHT
                    elif align in ("center", "centerContinuous"):
                        pango_align = Pango.Alignment.CENTER
                    elif isinstance(align_val, bool):
                        pango_align = Pango.Alignment.CENTER
                    elif isinstance(align_val, (int, float, datetime, date, time)):
                        pango_align = Pango.Alignment.RIGHT
                    else:
                        pango_align = Pango.Alignment.LEFT
                    layout.set_alignment(pango_align)

                    _, logical = layout.get_pixel_extents()
                    ty = sy + max(CELL_PADDING, (h_doc - logical.height) / 2)
                    cr.move_to(sx + CELL_PADDING, ty)
                    PangoCairo.show_layout(cr, layout)

        self._stroke_selection_border(cr, font_desc, sel_t, sel_l, sel_b, sel_r)
        cr.restore()

        cr.save()
        cr.rectangle(hcw, 0, vw - hcw, hrh)
        cr.clip()
        for col in range(first_col, last_col + 1):
            x_doc = self._col_x(col)
            w = cws[col - 1]
            sx = x_doc - self.scroll_x
            active = sel_l <= col <= sel_r
            bg = COLOR_HEADER_BG_ACTIVE if active else COLOR_HEADER_BG
            cr.set_source_rgb(*bg)
            cr.rectangle(sx, 0, w, hrh)
            cr.fill()
            cr.set_source_rgb(*COLOR_GRID)
            cr.set_line_width(1)
            cr.rectangle(sx + 0.5, 0.5, w, hrh)
            cr.stroke()
            cr.set_source_rgb(*COLOR_HEADER_TEXT)
            layout = PangoCairo.create_layout(cr)
            layout.set_font_description(header_font)
            layout.set_text(get_column_letter(col), -1)
            layout.set_width(w * Pango.SCALE)
            layout.set_alignment(Pango.Alignment.CENTER)
            _, logical = layout.get_pixel_extents()
            cr.move_to(sx, (hrh - logical.height) / 2)
            PangoCairo.show_layout(cr, layout)
        cr.restore()

        cr.save()
        cr.rectangle(0, hrh, hcw, vh - hrh)
        cr.clip()
        for row in range(first_row, last_row + 1):
            y_doc = self._row_y(row)
            h = rhs[row - 1]
            sy = y_doc - self.scroll_y
            active = sel_t <= row <= sel_b
            bg = COLOR_HEADER_BG_ACTIVE if active else COLOR_HEADER_BG
            cr.set_source_rgb(*bg)
            cr.rectangle(0, sy, hcw, h)
            cr.fill()
            cr.set_source_rgb(*COLOR_GRID)
            cr.set_line_width(1)
            cr.rectangle(0.5, sy + 0.5, hcw, h)
            cr.stroke()
            cr.set_source_rgb(*COLOR_HEADER_TEXT)
            layout = PangoCairo.create_layout(cr)
            layout.set_font_description(header_font)
            layout.set_text(str(row), -1)
            layout.set_width(hcw * Pango.SCALE)
            layout.set_alignment(Pango.Alignment.CENTER)
            _, logical = layout.get_pixel_extents()
            cr.move_to(0, sy + (h - logical.height) / 2)
            PangoCairo.show_layout(cr, layout)
        cr.restore()

        cr.set_source_rgb(*COLOR_HEADER_BG)
        cr.rectangle(0, 0, hcw, hrh)
        cr.fill()
        cr.set_source_rgb(*COLOR_GRID)
        cr.set_line_width(1)
        cr.rectangle(0.5, 0.5, hcw, hrh)
        cr.stroke()
        return False

    def _stroke_selection_border(self, cr, font_desc, sel_t, sel_l, sel_b, sel_r):
        rb, cb = sel_b, sel_r
        for (r, c), (nr, nc) in self.merged_anchor.items():
            if sel_t <= r <= sel_b and sel_l <= c <= sel_r:
                rb = max(rb, r + nr - 1)
                cb = max(cb, c + nc - 1)
        x_doc = self._col_x(sel_l)
        y_doc = self._row_y(sel_t)
        right_doc = self._col_x(cb + 1)
        bottom_doc = self._row_y(rb + 1)
        sx = x_doc - self.scroll_x
        sy = y_doc - self.scroll_y
        sw = right_doc - x_doc
        sh = bottom_doc - y_doc
        cr.set_source_rgb(*COLOR_SEL_BORDER)
        cr.set_line_width(2)
        cr.rectangle(sx + 0.5, sy + 0.5, sw - 1, sh - 1)
        cr.stroke()

    def _sel_rect(self):
        r1, c1 = self.sel_anchor
        r2, c2 = self.sel_cursor
        return min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2)

    def _on_button_press(self, widget, event):
        self.drawing_area.grab_focus()
        if event.button != 1:
            return False
        x_doc = event.x + self.scroll_x
        y_doc = event.y + self.scroll_y
        col = self._col_at_x(x_doc)
        row = self._row_at_y(y_doc)

        if row == 0 and col == 0:
            self.sel_anchor = (1, 1)
            self.sel_cursor = (self.max_row, self.max_col)
            self.dragging_sel = False
        elif row == 0:
            self.sel_anchor = (1, col)
            self.sel_cursor = (self.max_row, col)
            self.dragging_sel = True
            self._drag_axis = "col"
        elif col == 0:
            self.sel_anchor = (row, 1)
            self.sel_cursor = (row, self.max_col)
            self.dragging_sel = True
            self._drag_axis = "row"
        else:
            if event.state & Gdk.ModifierType.SHIFT_MASK:
                self.sel_cursor = (row, col)
            else:
                self.sel_anchor = (row, col)
                self.sel_cursor = (row, col)
            self.dragging_sel = True
            self._drag_axis = "cell"

        self._notify_selection()
        self.drawing_area.queue_draw()
        return True

    def _on_motion(self, widget, event):
        if not self.dragging_sel:
            return False
        x_doc = event.x + self.scroll_x
        y_doc = event.y + self.scroll_y
        col = max(1, min(self.max_col, self._col_at_x(x_doc) or 1))
        row = max(1, min(self.max_row, self._row_at_y(y_doc) or 1))
        if self._drag_axis == "col":
            self.sel_cursor = (self.max_row, col)
        elif self._drag_axis == "row":
            self.sel_cursor = (row, self.max_col)
        else:
            self.sel_cursor = (row, col)
        self._notify_selection()
        self.drawing_area.queue_draw()
        return True

    def _on_button_release(self, widget, event):
        self.dragging_sel = False
        return True

    def _on_scroll(self, widget, event):
        if event.state & Gdk.ModifierType.CONTROL_MASK:
            old_zoom = self.zoom
            if event.direction == Gdk.ScrollDirection.SMOOTH:
                _, dy = _scroll_deltas(event)
                if dy < 0:
                    self.zoom *= 1.1 ** abs(dy)
                elif dy > 0:
                    self.zoom /= 1.1 ** abs(dy)
            elif event.direction == Gdk.ScrollDirection.UP:
                self.zoom *= 1.1
            elif event.direction == Gdk.ScrollDirection.DOWN:
                self.zoom /= 1.1
            else:
                return False
            self.zoom = max(MIN_ZOOM, min(MAX_ZOOM, self.zoom))
            if abs(self.zoom - old_zoom) < 1e-6:
                return True

            ratio = self.zoom / old_zoom
            self.scroll_x = int((self.scroll_x + event.x) * ratio - event.x)
            self.scroll_y = int((self.scroll_y + event.y) * ratio - event.y)
            self._update_adjustments()
            max_hscroll = int(self.hadj.get_upper() - self.hadj.get_page_size())
            max_vscroll = int(self.vadj.get_upper() - self.vadj.get_page_size())
            self.scroll_x = max(0, min(self.scroll_x, max_hscroll))
            self.scroll_y = max(0, min(self.scroll_y, max_vscroll))
            self.hadj.set_value(self.scroll_x)
            self.vadj.set_value(self.scroll_y)
            if self.on_zoom_changed:
                self.on_zoom_changed(self.zoom)
            self.drawing_area.queue_draw()
            return True

        step = 40
        if event.direction == Gdk.ScrollDirection.SMOOTH:
            dx, dy = _scroll_deltas(event)
            if event.state & Gdk.ModifierType.SHIFT_MASK:
                self.hadj.set_value(self.hadj.get_value() + dy * step)
            else:
                self.vadj.set_value(self.vadj.get_value() + dy * step)
                if dx:
                    self.hadj.set_value(self.hadj.get_value() + dx * step)
            return True
        if event.direction == Gdk.ScrollDirection.UP:
            self.vadj.set_value(self.vadj.get_value() - step)
        elif event.direction == Gdk.ScrollDirection.DOWN:
            self.vadj.set_value(self.vadj.get_value() + step)
        elif event.direction == Gdk.ScrollDirection.LEFT:
            self.hadj.set_value(self.hadj.get_value() - step)
        elif event.direction == Gdk.ScrollDirection.RIGHT:
            self.hadj.set_value(self.hadj.get_value() + step)
        return True

    def _on_key_press(self, widget, event):
        extend = bool(event.state & Gdk.ModifierType.SHIFT_MASK)
        ctrl = bool(event.state & Gdk.ModifierType.CONTROL_MASK)
        r, c = self.sel_cursor
        moved = True
        if event.keyval in (Gdk.KEY_Up, Gdk.KEY_KP_Up):
            r = max(1, r - 1)
        elif event.keyval in (Gdk.KEY_Down, Gdk.KEY_KP_Down):
            r = min(self.max_row, r + 1)
        elif event.keyval in (Gdk.KEY_Left, Gdk.KEY_KP_Left):
            c = max(1, c - 1)
        elif event.keyval in (Gdk.KEY_Right, Gdk.KEY_KP_Right):
            c = min(self.max_col, c + 1)
        elif event.keyval == Gdk.KEY_Home:
            c = 1
            if ctrl:
                r = 1
        elif event.keyval == Gdk.KEY_End:
            c = self.max_col
            if ctrl:
                r = self.max_row
        elif event.keyval in (Gdk.KEY_Page_Up, Gdk.KEY_Page_Down):
            alloc = self.drawing_area.get_allocation()
            page = max(1, (alloc.height - self.header_row_height) // DEFAULT_ROW_HEIGHT)
            if event.keyval == Gdk.KEY_Page_Up:
                r = max(1, r - page)
            else:
                r = min(self.max_row, r + page)
        else:
            moved = False

        if moved:
            self.sel_cursor = (r, c)
            if not extend:
                self.sel_anchor = (r, c)
            self._scroll_to_cell(r, c)
            self._notify_selection()
            self.drawing_area.queue_draw()
            return True
        return False

    def _scroll_to_cell(self, row, col):
        alloc = self.drawing_area.get_allocation()
        x_doc = self._col_x(col)
        y_doc = self._row_y(row)
        w = self.col_widths[col - 1]
        h = self.row_heights[row - 1]
        if x_doc < self.scroll_x + self.header_col_width:
            self.hadj.set_value(max(0, x_doc - self.header_col_width))
        elif x_doc + w > self.scroll_x + alloc.width:
            self.hadj.set_value(x_doc + w - alloc.width)
        if y_doc < self.scroll_y + self.header_row_height:
            self.vadj.set_value(max(0, y_doc - self.header_row_height))
        elif y_doc + h > self.scroll_y + alloc.height:
            self.vadj.set_value(y_doc + h - alloc.height)

    def _notify_selection(self):
        if self.on_selection_changed:
            self.on_selection_changed(self)

    # ----- formula evaluation ---------------------------------------------

    def resolve_display_value(self, row, col):
        """Return the value to show for (row, col): a scalar, FormulaErrorValue,
        or FORMULA_PENDING. Never raises. Only reads caches; triggers async
        eval when a visible formula cell isn't cached yet."""
        cached = self.values.get((row, col))
        if cached is not None:
            return cached
        f = self.formulas.get((row, col))
        if f is None:
            return None
        if (row, col) in self.formula_cache:
            return self.formula_cache[(row, col)]
        if self.evaluator is not None:
            self.evaluator.enqueue(self, row, col)
        return FORMULA_PENDING

    def formula_for_cell(self, row, col):
        if (row, col) in self.merged_covered:
            row, col = self.merged_covered[(row, col)]
        return self.formulas.get((row, col))

    def _display_value_at(self, row, col):
        if (row, col) in self.merged_covered:
            row, col = self.merged_covered[(row, col)]
        return self.resolve_display_value(row, col)

    def _format_for_copy(self, v):
        if v is FORMULA_PENDING:
            return ""
        if isinstance(v, FormulaErrorValue):
            return "#ERR"
        return format_cell_value(v)

    def get_selection_text(self):
        t, l, b, r = self._sel_rect()
        lines = []
        for row in range(t, b + 1):
            fields = []
            for col in range(l, r + 1):
                v = self._display_value_at(row, col)
                fields.append(self._format_for_copy(v))
            lines.append("\t".join(fields))
        return "\n".join(lines)

    def selection_info(self):
        t, l, b, r = self._sel_rect()
        if t == b and l == r:
            return f"{get_column_letter(l)}{t}"
        return f"{get_column_letter(l)}{t}:{get_column_letter(r)}{b}"

    def selection_stats(self):
        t, l, b, r = self._sel_rect()
        count = 0
        num_count = 0
        total = 0.0
        vmin = None
        vmax = None
        for row in range(t, b + 1):
            for col in range(l, r + 1):
                if (row, col) in self.merged_covered:
                    continue
                v = self.resolve_display_value(row, col)
                if v is None or v == "" or v is FORMULA_PENDING:
                    continue
                if isinstance(v, FormulaErrorValue):
                    continue
                count += 1
                if isinstance(v, bool):
                    continue
                if isinstance(v, (int, float)):
                    num_count += 1
                    total += v
                    vmin = v if vmin is None or v < vmin else vmin
                    vmax = v if vmax is None or v > vmax else vmax
        return {
            "rows": b - t + 1,
            "cols": r - l + 1,
            "count": count,
            "num_count": num_count,
            "sum": total if num_count else None,
            "avg": total / num_count if num_count else None,
            "min": vmin,
            "max": vmax,
        }

    def anchor_value_text(self):
        t, l, _, _ = self._sel_rect()
        v = self._display_value_at(t, l)
        if v is FORMULA_PENDING:
            return "…"
        if isinstance(v, FormulaErrorValue):
            return f"#ERR: {v.msg}"
        return format_cell_value(v)

    def set_zoom(self, z):
        old = self.zoom
        self.zoom = max(MIN_ZOOM, min(MAX_ZOOM, z))
        if abs(self.zoom - old) < 1e-6:
            return
        self._update_adjustments()
        self.drawing_area.queue_draw()
        if self.on_zoom_changed:
            self.on_zoom_changed(self.zoom)


class QuickCellApp:
    def __init__(self):
        self.window = Gtk.Window(title="QuickCell")
        self.window.set_default_size(1100, 720)
        self.window.connect("destroy", Gtk.main_quit)
        self.window.connect("key-press-event", self._on_window_key)

        self.filepath = None
        self.wb = None
        self.evaluator = None
        self.sheet_views = []

        open_btn = Gtk.Button(label="📂 Open")
        open_btn.connect("clicked", lambda *_: self.open_dialog())
        copy_btn = Gtk.Button(label="📄 Copy")
        copy_btn.connect("clicked", lambda *_: self.copy_selection())
        zout_btn = Gtk.Button(label="🔍−")
        zout_btn.connect("clicked", lambda *_: self.zoom_delta(1 / 1.1))
        zreset_btn = Gtk.Button(label="1:1")
        zreset_btn.connect("clicked", lambda *_: self.zoom_set(1.0))
        zin_btn = Gtk.Button(label="🔍+")
        zin_btn.connect("clicked", lambda *_: self.zoom_delta(1.1))
        help_btn = Gtk.Button(label="❓ Help")
        help_btn.connect("clicked", lambda *_: self.show_help())

        def vsep():
            return Gtk.Separator(orientation=Gtk.Orientation.VERTICAL)

        hbox = Gtk.Box(spacing=6)
        hbox.pack_start(open_btn, False, False, 0)
        hbox.pack_start(copy_btn, False, False, 0)
        hbox.pack_start(vsep(), False, False, 0)
        hbox.pack_start(zout_btn, False, False, 0)
        hbox.pack_start(zreset_btn, False, False, 0)
        hbox.pack_start(zin_btn, False, False, 0)
        hbox.pack_start(vsep(), False, False, 0)
        hbox.pack_start(help_btn, False, False, 0)

        self.formula_cell_label = Gtk.Label(label="")
        self.formula_cell_label.set_xalign(0.5)
        self.formula_cell_label.set_size_request(80, -1)
        fx_label = Gtk.Label()
        fx_label.set_markup("<i>fx</i>")
        fx_label.set_margin_start(4)
        fx_label.set_margin_end(4)
        self.formula_entry = Gtk.Entry()
        self.formula_entry.set_editable(False)
        self.formula_entry.set_can_focus(True)
        self.formula_entry.set_hexpand(True)
        formula_box = Gtk.Box(spacing=4)
        formula_box.pack_start(self.formula_cell_label, False, False, 0)
        formula_box.pack_start(vsep(), False, False, 0)
        formula_box.pack_start(fx_label, False, False, 0)
        formula_box.pack_start(self.formula_entry, True, True, 0)

        self.notebook = Gtk.Notebook()
        self.notebook.set_scrollable(True)
        self.notebook.set_tab_pos(Gtk.PositionType.BOTTOM)
        self.notebook.connect("switch-page", self._on_switch_page)

        self.status_cell = Gtk.Label(label="")
        self.status_cell.set_xalign(0)
        self.status_cell.set_size_request(110, -1)
        self.status_value = Gtk.Label(label="")
        self.status_value.set_xalign(0)
        self.status_value.set_ellipsize(Pango.EllipsizeMode.END)
        self.status_zoom = Gtk.Label(label="100%")
        self.status_zoom.set_xalign(1)
        self.status_zoom.set_size_request(60, -1)

        status_box = Gtk.Box(spacing=12)
        status_box.set_margin_start(8)
        status_box.set_margin_end(8)
        status_box.set_margin_top(2)
        status_box.set_margin_bottom(2)
        status_box.pack_start(self.status_cell, False, False, 0)
        status_box.pack_start(self.status_value, True, True, 0)
        status_box.pack_start(self.status_zoom, False, False, 0)

        self.toast_label = Gtk.Label()
        css = Gtk.CssProvider()
        css.load_from_data(
            b".toast { background-color: rgba(0,0,0,0.8); color: white; "
            b"padding: 10px 20px; border-radius: 5px; }"
        )
        self.toast_label.get_style_context().add_provider(
            css, Gtk.STYLE_PROVIDER_PRIORITY_APPLICATION
        )
        self.toast_label.get_style_context().add_class("toast")
        self.toast_label.set_halign(Gtk.Align.CENTER)
        self.toast_label.set_valign(Gtk.Align.START)
        self.toast_label.set_margin_top(10)
        self.toast_label.set_no_show_all(True)
        self._toast_timer = None

        self.loading_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=8)
        self.loading_box.set_halign(Gtk.Align.CENTER)
        self.loading_box.set_valign(Gtk.Align.CENTER)
        self.loading_spinner = Gtk.Spinner()
        self.loading_spinner.set_size_request(48, 48)
        self.loading_label = Gtk.Label(label="Loading…")
        self.loading_detail = Gtk.Label(label="")
        detail_css = Gtk.CssProvider()
        detail_css.load_from_data(
            b".loading-detail { color: rgba(0,0,0,0.55); font-size: smaller; }"
        )
        self.loading_detail.get_style_context().add_provider(
            detail_css, Gtk.STYLE_PROVIDER_PRIORITY_APPLICATION
        )
        self.loading_detail.get_style_context().add_class("loading-detail")
        self.loading_box.pack_start(self.loading_spinner, False, False, 0)
        self.loading_box.pack_start(self.loading_label, False, False, 0)
        self.loading_box.pack_start(self.loading_detail, False, False, 0)
        self.loading_box.set_no_show_all(True)

        self.overlay = Gtk.Overlay()
        self.overlay.add_overlay(self.toast_label)
        self.overlay.add_overlay(self.loading_box)

        vbox = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=4)
        vbox.set_margin_top(4)
        vbox.set_margin_bottom(4)
        vbox.set_margin_start(4)
        vbox.set_margin_end(4)
        vbox.pack_start(hbox, False, False, 0)
        vbox.pack_start(formula_box, False, False, 0)
        vbox.pack_start(self.notebook, True, True, 0)
        vbox.pack_start(status_box, False, False, 0)

        self.overlay.add(vbox)
        self.window.add(self.overlay)
        self.window.show_all()

        self._show_empty_state()

    def _show_empty_state(self):
        placeholder = Gtk.Label(
            label="Open an .xlsx file (Ctrl+O, or pass a path on the command line)"
        )
        placeholder.set_vexpand(True)
        placeholder.set_hexpand(True)
        self.notebook.append_page(placeholder, Gtk.Label(label="(no file)"))
        self.notebook.show_all()
        self.formula_cell_label.set_text("")
        self.formula_entry.set_text("")
        self.status_cell.set_text("")
        self.status_value.set_text("")

    def show_toast(self, message):
        if self._toast_timer:
            GLib.source_remove(self._toast_timer)
        self.toast_label.set_text(message)
        self.toast_label.show()
        self._toast_timer = GLib.timeout_add(2000, self._hide_toast)

    def _hide_toast(self):
        self._toast_timer = None
        self.toast_label.hide()
        return False

    def load_file(self, path):
        if not os.path.exists(path):
            self.show_toast(f"File not found: {os.path.basename(path)}")
            return False

        while self.notebook.get_n_pages() > 0:
            self.notebook.remove_page(-1)
        self.sheet_views = []
        self.wb = None
        self.filepath = path
        self.window.set_title(f"QuickCell — Loading {os.path.basename(path)}…")
        self._show_loading(path)

        t = threading.Thread(target=self._load_thread, args=(path,), daemon=True)
        t.start()
        return True

    def _load_thread(self, path):
        try:
            GLib.idle_add(self._set_loading_detail, "Reading workbook…")
            wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
            GLib.idle_add(self._set_loading_detail, "Reading formulas…")
            try:
                wb_f = openpyxl.load_workbook(path, data_only=False, read_only=False)
            except Exception:
                wb_f = None

            prebuilt = []
            names = wb.sheetnames
            for i, name in enumerate(names, 1):
                GLib.idle_add(
                    self._set_loading_detail,
                    f"Indexing sheet {i}/{len(names)}: {name}…",
                )
                ws = wb[name]
                values = {}
                for row in ws.iter_rows():
                    for cell in row:
                        v = cell.value
                        if v is not None:
                            values[(cell.row, cell.column)] = v
                formulas = {}
                ws_f = None
                if wb_f is not None and name in wb_f.sheetnames:
                    ws_f = wb_f[name]
                    for row in ws_f.iter_rows():
                        for cell in row:
                            v = cell.value
                            if isinstance(v, str) and v.startswith("="):
                                formulas[(cell.row, cell.column)] = v
                prebuilt.append((name, ws, ws_f, values, formulas))
            GLib.idle_add(self._set_loading_detail, "Building views…")
            GLib.idle_add(self._on_load_done, path, wb, wb_f, prebuilt, None)
        except Exception as e:
            GLib.idle_add(self._on_load_done, path, None, None, None, str(e))

    def _on_load_done(self, path, wb, wb_f, prebuilt, err):
        self._hide_loading()
        if err is not None or wb is None:
            self.show_toast(f"Cannot open: {err}")
            self.window.set_title("QuickCell")
            self._show_empty_state()
            return False

        self.wb = wb
        self.evaluator = WorkbookEvaluator()
        for name, ws, ws_f, values, formulas in prebuilt:
            view = SheetView(
                ws,
                sheet_name=name,
                ws_formulas=ws_f,
                evaluator=self.evaluator,
                prebuilt_values=values,
                prebuilt_formulas=formulas,
                on_selection_changed=self._on_selection_changed,
                on_zoom_changed=self._on_zoom_changed,
            )
            self.sheet_views.append(view)
            self.notebook.append_page(view, Gtk.Label(label=name))

        self.notebook.show_all()
        self.window.set_title(f"QuickCell — {os.path.basename(path)}")
        self.notebook.set_current_page(0)
        if self.sheet_views:
            v = self.sheet_views[0]
            GLib.idle_add(v.drawing_area.grab_focus)
            self._on_selection_changed(v)
            self._on_zoom_changed(v.zoom)
        return False

    def _show_loading(self, path):
        self.loading_label.set_text(f"Loading {os.path.basename(path)}…")
        self.loading_detail.set_text("")
        self.loading_box.show()
        self.loading_spinner.show()
        self.loading_label.show()
        self.loading_detail.show()
        self.loading_spinner.start()

    def _hide_loading(self):
        self.loading_spinner.stop()
        self.loading_box.hide()
        self.loading_detail.set_text("")

    def _set_loading_detail(self, text):
        self.loading_detail.set_text(text)
        return False

    def open_dialog(self):
        dialog = Gtk.FileChooserDialog(
            title="Open spreadsheet",
            parent=self.window,
            action=Gtk.FileChooserAction.OPEN,
        )
        dialog.add_button("Cancel", Gtk.ResponseType.CANCEL)
        dialog.add_button("Open", Gtk.ResponseType.OK)
        flt = Gtk.FileFilter()
        flt.set_name("Excel files (xlsx, xlsm)")
        for ext in ("xlsx", "xlsm"):
            flt.add_pattern(_case_insensitive_glob(ext))
        dialog.add_filter(flt)
        any_flt = Gtk.FileFilter()
        any_flt.set_name("All files")
        any_flt.add_pattern("*")
        dialog.add_filter(any_flt)
        resp = dialog.run()
        path = dialog.get_filename() if resp == Gtk.ResponseType.OK else None
        dialog.destroy()
        if path:
            self.load_file(path)

    def current_view(self):
        page = self.notebook.get_current_page()
        if 0 <= page < len(self.sheet_views):
            return self.sheet_views[page]
        return None

    def _on_switch_page(self, notebook, page, page_num):
        if 0 <= page_num < len(self.sheet_views):
            v = self.sheet_views[page_num]
            self._on_selection_changed(v)
            self._on_zoom_changed(v.zoom)
            GLib.idle_add(v.drawing_area.grab_focus)

    def _on_selection_changed(self, view):
        if view is not self.current_view():
            return
        self.status_cell.set_text(view.selection_info())
        self._update_formula_bar(view)
        stats = view.selection_stats()
        if stats["rows"] == 1 and stats["cols"] == 1:
            self.status_value.set_text(view.anchor_value_text())
            return
        parts = [f"{stats['rows']}R × {stats['cols']}C"]
        parts.append(f"Count: {stats['count']}")
        if stats["num_count"]:
            parts.append(f"Sum: {_fmt_stat(stats['sum'])}")
            parts.append(f"Avg: {_fmt_stat(stats['avg'])}")
            parts.append(f"Min: {_fmt_stat(stats['min'])}")
            parts.append(f"Max: {_fmt_stat(stats['max'])}")
        self.status_value.set_text("   ".join(parts))

    def _on_zoom_changed(self, zoom):
        self.status_zoom.set_text(f"{int(round(zoom * 100))}%")

    def _update_formula_bar(self, view):
        t, l, _, _ = view._sel_rect()
        self.formula_cell_label.set_text(f"{get_column_letter(l)}{t}")
        f = view.formula_for_cell(t, l)
        if f:
            self.formula_entry.set_text(f)
        else:
            self.formula_entry.set_text(view.anchor_value_text())

    def copy_selection(self):
        view = self.current_view()
        if view is None:
            return
        text = view.get_selection_text()
        clipboard = Gtk.Clipboard.get(Gdk.SELECTION_CLIPBOARD)
        clipboard.set_text(text, -1)
        self.show_toast(f"✓ Copied {view.selection_info()}")

    def zoom_delta(self, factor):
        view = self.current_view()
        if view is None:
            return
        view.set_zoom(view.zoom * factor)

    def zoom_set(self, z):
        view = self.current_view()
        if view is None:
            return
        view.set_zoom(z)

    def _on_window_key(self, widget, event):
        ctrl = bool(event.state & Gdk.ModifierType.CONTROL_MASK)
        if not ctrl:
            return False
        k = event.keyval
        focus = self.window.get_focus()
        in_entry = focus is self.formula_entry
        if k == Gdk.KEY_a:
            if in_entry:
                self.formula_entry.select_region(0, -1)
            else:
                view = self.current_view()
                if view is not None:
                    view.sel_anchor = (1, 1)
                    view.sel_cursor = (view.max_row, view.max_col)
                    view._notify_selection()
                    view.drawing_area.queue_draw()
            return True
        if k == Gdk.KEY_c:
            if in_entry:
                return False  # let Gtk.Entry copy selected text
            self.copy_selection()
            return True
        if k == Gdk.KEY_o:
            self.open_dialog()
            return True
        if k in (Gdk.KEY_plus, Gdk.KEY_equal, Gdk.KEY_KP_Add):
            self.zoom_delta(1.1)
            return True
        if k in (Gdk.KEY_minus, Gdk.KEY_KP_Subtract):
            self.zoom_delta(1 / 1.1)
            return True
        if k == Gdk.KEY_0:
            self.zoom_set(1.0)
            return True
        if k == Gdk.KEY_Page_Down:
            self.notebook.next_page()
            return True
        if k == Gdk.KEY_Page_Up:
            self.notebook.prev_page()
            return True
        return False

    def show_help(self):
        dialog = Gtk.Dialog(title="Help", parent=self.window, flags=0)
        dialog.add_button("Close", Gtk.ResponseType.CLOSE)
        text = Gtk.Label()
        text.set_markup(
            f"""<b>QuickCell v{VERSION}</b>

A minimal read-only viewer for .xlsx files.

<b>Mouse:</b>
• Click / drag: select a cell or range
• Shift+click: extend selection
• Click column/row header: select whole column/row
• Click top-left corner: select all

<b>Keyboard:</b>
• Arrow keys / Home / End / PgUp / PgDn: navigate
• Shift+arrows: extend selection
• Ctrl+A: select all cells (or all text in the formula bar)
• Ctrl+Home / Ctrl+End: jump to A1 / last cell
• Ctrl+C: copy selection (tab-separated)
• Ctrl+O: open file
• Ctrl+scroll / Ctrl++ / Ctrl+− / Ctrl+0: zoom
• Ctrl+PgUp / Ctrl+PgDn: switch sheet

<b>Formula bar:</b>
• Shows the raw formula (or value) of the selected cell.
• Read-only; click into it and Ctrl+A / Ctrl+C to copy the formula.

<b>Notes:</b>
• Files are opened read-only; nothing is ever written back.
• Formula results come from the cached value stored in the file.
  For cells without a cached value, QuickCell evaluates formulas
  in the background, including cross-sheet refs in the same file.
  Cells being evaluated show a yellow tint; cells whose formulas
  can't be evaluated show "Error" in red.

<b>GitHub:</b> github.com/yonie/quickcell

<b>License:</b> MIT"""
        )
        text.set_margin_start(15)
        text.set_margin_end(15)
        text.set_margin_top(15)
        text.set_margin_bottom(15)
        dialog.get_content_area().add(text)
        dialog.show_all()
        dialog.run()
        dialog.destroy()

    def _load_initial_file(self, path):
        self.load_file(path)
        return GLib.SOURCE_REMOVE


def main():
    app = QuickCellApp()
    if len(sys.argv) > 1:
        GLib.idle_add(app._load_initial_file, sys.argv[1])
    Gtk.main()


if __name__ == "__main__":
    main()
